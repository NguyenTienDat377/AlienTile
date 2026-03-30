from docplex.mp.model import Model
import gurobipy as gp

import json
import math
import sys
import time
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


def export_to_excel(N, c, target, result, elapsed, variant, path="results.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def write_matrix(matrix, start_row, start_col, title, color_map=None):
        ws.cell(row=start_row, column=start_col, value=title).font = Font(bold=True)
        for r in range(N):
            for k in range(N):
                val = matrix[r][k]
                cell = ws.cell(row=start_row + 1 + r, column=start_col + k, value=val)
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
                if color_map and val in color_map:
                    cell.fill = PatternFill("solid", fgColor=color_map[val])

    ws["A1"] = "N"; ws["B1"] = N
    ws["A2"] = "c"; ws["B2"] = c
    ws["A3"] = "Variant"; ws["B3"] = variant
    ws["A4"] = "Time (s)"; ws["B4"] = round(elapsed, 4)
    ws["A5"] = "Total clicks"; ws["B5"] = result["total_clicks"] if result else "N/A"
    ws["A6"] = "Verification"; ws["B6"] = "PASS" if result and verify_solution(N, c, result["T"] if result["T"] else target, result["X"]) else "FAIL"

    palette = ["FFFFFF", "4472C4", "ED7D31", "A9D18E", "FF0000",
               "FFFF00", "9B59B6", "1ABC9C", "E74C3C", "F39C12"]
    color_map = {i: palette[i % len(palette)] for i in range(c)}

    write_matrix(target, start_row=8, start_col=1, title="Target", color_map=color_map)
    if result:
        write_matrix(result["X"], start_row=8, start_col=N + 3, title="Click matrix X")
        if result["T"] is not None:
            write_matrix(result["T"], start_row=8, start_col=2 * N + 5, title="Hardest target T", color_map=color_map)

    wb.save(path)
    print(f"Results exported to {path}")


def load_instance(path: str):
    with open(path) as f:
        data = json.load(f)
    return data["N"], data["c"], data["target"]


def print_board(board, title="Board"):
    print(f"{title}:")
    for row in board:
        print("  " + " ".join(f"{cell}" for cell in row))
    print()


def verify_solution(N, c, target, X):
    for r in range(N):
        for k in range(N):
            sigma = sum(X[r][j] for j in range(N)) \
                  + sum(X[i][k] for i in range(N)) \
                  - X[r][k]
            if sigma % c != target[r][k]:
                print(f"  MISMATCH at ({r},{k}): "
                      f"sigma={sigma}, sigma mod {c}={sigma % c}, "
                      f"target={target[r][k]}")
                return False
    return True

class AlienTilesILP:
    def __init__(self, N, c, target=None):
        self.N = N
        self.c = c
        self.target = target
        self.x_ij = [[None] * N for _ in range(N)]
        self.q_rk = [[None] * N for _ in range(N)]
        self.t_rk = [[None] * N for _ in range(N)]
        self.mdl = Model(name="AlienTilesILP")

    def _encode_variables(self) -> None:
        upper_bound = math.floor((2 * self.N - 1) * (self.c - 1) / self.c)
        for i in range(self.N):
            for j in range(self.N):
                self.x_ij[i][j] = self.mdl.integer_var(
                    name=f"x_{i}_{j}", lb=0, ub=self.c - 1
                )
        for r in range(self.N):
            for k in range(self.N):
                self.q_rk[r][k] = self.mdl.integer_var(
                    name=f"q_{r}_{k}", lb=0, ub=upper_bound
                )

    def _encode_constraints(self, target_vars=None) -> None:
        t = target_vars if target_vars else self.target
        for r in range(self.N):
            for k in range(self.N):
                sigma = (self.mdl.sum(self.x_ij[r][j] for j in range(self.N))
                         + self.mdl.sum(self.x_ij[i][k] for i in range(self.N))
                         - self.x_ij[r][k])
                self.mdl.add_constraint(
                    sigma - self.c * self.q_rk[r][k] == t[r][k],
                    ctname=f"mod_eq_{r}_{k}"
                )

    def _encode_symmetry_breaking(self) -> None:
        N = self.N

        for i in range(N - 1):
            self.mdl.add_constraint(
                self.mdl.sum(self.x_ij[i][j] for j in range(N))
                <= self.mdl.sum(self.x_ij[i + 1][j] for j in range(N)),
                ctname=f"row_sym_{i}"
            )
        for j in range(N - 1):
            self.mdl.add_constraint(
                self.mdl.sum(self.x_ij[i][j] for i in range(N))
                <= self.mdl.sum(self.x_ij[i][j + 1] for i in range(N)),
                ctname=f"col_sym_{j}"
            )
        self.mdl.add_constraint(self.x_ij[0][1] <= self.x_ij[1][0],
                                ctname="diag_sym")

    def build_variant1(self, symmetry_breaking=False):
        self._encode_variables()
        self._encode_constraints()
        if symmetry_breaking:
            self._encode_symmetry_breaking()

    def build_variant2(self, symmetry_breaking=False):
        self._encode_variables()
        self._encode_constraints()
        if symmetry_breaking:
            self._encode_symmetry_breaking()
        self.mdl.minimize(
            self.mdl.sum(self.x_ij[i][j] for i in range(self.N) for j in range(self.N))
        )

    def build_variant3(self, symmetry_breaking=False):
        self._encode_variables()
        # Make target a decision variable
        for r in range(self.N):
            for k in range(self.N):
                self.t_rk[r][k] = self.mdl.integer_var(
                    name=f"t_{r}_{k}", lb=0, ub=self.c - 1
                )
        self._encode_constraints(target_vars=self.t_rk)

        self.mdl.add_constraint(
            self.mdl.sum(self.t_rk[r][k]
                         for r in range(self.N) for k in range(self.N)) >= 1,
            ctname="non_trivial_target"
        )
        if symmetry_breaking:
            self._encode_symmetry_breaking()

        self.mdl.maximize(
            self.mdl.sum(self.x_ij[i][j] for i in range(self.N) for j in range(self.N))
        )

    def solve(self):
        solution = self.mdl.solve()
        if solution is None:
            return None
        X = [[int(self.x_ij[i][j].solution_value)
              for j in range(self.N)] for i in range(self.N)]
        T = None
        if self.t_rk[0][0] is not None:
            T = [[int(self.t_rk[r][k].solution_value)
                  for k in range(self.N)] for r in range(self.N)]
        total = sum(X[i][j] for i in range(self.N) for j in range(self.N))
        return {"X": X, "T": T, "total_clicks": total}


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(f"Usage: python {sys.argv[0]} <instance.json> <variant> [output.xlsx]")
        print("  variant: 1 (feasibility) | 2 (min clicks) | 3 (hardest puzzle)")
        sys.exit(1)

    path = sys.argv[1]
    variant = int(sys.argv[2]) if len(sys.argv) > 2 else 2
    xlsx_path = sys.argv[3] if len(sys.argv) > 3 else "results.xlsx"

    N, c, target = load_instance(path)
    print(f"Instance: N={N}, c={c}")
    print_board(target, "Target")

    ilp = AlienTilesILP(N, c, target)

    print(f"--- Variant {variant} ---")
    t0 = time.time()

    if variant == 1:
        ilp.build_variant1()
    elif variant == 2:
        ilp.build_variant2()
    elif variant == 3:
        ilp.build_variant3()
    else:
        print(f"Unknown variant: {variant}")
        sys.exit(1)

    result = ilp.solve()
    elapsed = time.time() - t0

    if result is None:
        print("No solution found (infeasible).")
    else:
        print_board(result["X"], "Click matrix X")
        print(f"Total clicks: {result['total_clicks']}")
        if result["T"] is not None:
            print_board(result["T"], "Hardest target T")
        ok = verify_solution(N, c,
                             result["T"] if result["T"] else target,
                             result["X"])
        print(f"Verification: {'PASS' if ok else 'FAIL'}")

    print(f"Time: {elapsed:.3f}s")
    export_to_excel(N, c, target, result, elapsed, variant, path=xlsx_path)
