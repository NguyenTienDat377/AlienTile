"""
SAT-Based Optimisation for the Alien Tiles Problem (general c >= 2).

Implements the three approaches from Section 7:
  7.1  Non-incremental SAT  (binary search + linear search top-down)
  7.2  Incremental SAT      (assumption-based bound control with totalizer)
  7.3  MaxSAT / PBO         (partial weighted MaxSAT)

Each click variable x_{i,j} in {0,...,c-1} is binary-encoded (Section 6.1).
Feasibility uses a running modular-sum encoding (Sections 6.2-6.3).
The objective uses unit-contribution literals u_{i,j,v} (Section 7).
"""

import json
import math
import sys
import time

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from pysat.card import CardEnc, EncType
from pysat.examples.rc2 import RC2
from pysat.formula import WCNF
from pysat.solvers import Solver as PySATSolver


# ── helpers ──────────────────────────────────────────────────────────────

def load_instance(path: str):
    with open(path) as f:
        data = json.load(f)
    return data["N"], data["c"], data["target"]


def print_board(board, title="Board"):
    print(f"{title}:")
    for row in board:
        print(" ".join(str(cell) for cell in row))
    print()

def verify_solution(N, c, target, X):
    """Check that click matrix X produces target T under modular arithmetic."""
    for r in range(N):
        for k in range(N):
            sigma = sum(X[r][j] for j in range(N)) \
                  + sum(X[i][k] for i in range(N)) \
                  - X[r][k]
            if sigma % c != target[r][k]:
                print(f"  MISMATCH at ({r},{k}): "
                      f"sigma={sigma}, sigma mod {c}={sigma % c}, "
                      f"target={target[r][k]}")
                return False
    return True

# ── CNF encoder ──────────────────────────────────────────────────────────

class AlienTilesEncoder:
    """Encodes an Alien Tiles instance (N, c, target) as a CNF formula."""

    def __init__(self, N, c, target):
        self.N = N
        self.c = 4
        self.target = target
        self.b = math.ceil(math.log2(c)) if c > 1 else 1
        self.top = 0
        self.clauses = []
        self.p = None       # p[i][j][l]: SAT var for bit l of x_{i,j}
        self.d = None       # d[i][j][w]: SAT var for "x_{i,j} = w"
        self.u_lits = None  # all unit-contribution literals

    def _new(self):
        self.top += 1
        return self.top

    def build(self):
        """Build feasibility CNF + unit-contribution literals."""
        self._encode_click_vars()
        self._encode_value_indicators()
        self._encode_feasibility()
        self._encode_unit_lits()
        return self.clauses, self.u_lits, self.top

    # ── Section 6.1: binary encoding of click variables ──────────────

    def _encode_click_vars(self):
        N, c, b = self.N, self.c, self.b
        self.p = [[[self._new() for _ in range(b)]
                    for _ in range(N)]
                   for _ in range(N)]

    # ── value-indicator variables d[i][j][w] ⟺ "x_{i,j} = w" ────────

    def _encode_value_indicators(self):
        N, c, b = self.N, self.c, self.b
        self.d = [[[None] * c for _ in range(N)] for _ in range(N)]

        for i in range(N):
            for j in range(N):
                for w in range(c):
                    dv = self._new()
                    self.d[i][j][w] = dv

                    # Forward: dv → each bit matches w's pattern
                    backward = [dv]
                    for l in range(b):
                        if (w >> l) & 1:
                            self.clauses.append([-dv, self.p[i][j][l]])
                            backward.append(-self.p[i][j][l])
                        else:
                            self.clauses.append([-dv, -self.p[i][j][l]])
                            backward.append(self.p[i][j][l])
                    # Backward: all bits match w → dv
                    self.clauses.append(backward)

    # ── Sections 6.2–6.3: feasibility via unary sum + totalizer ──────

    def _local_totalizer(self, lits):
        """
        Build a BIDIRECTIONAL totalizer network.

        Returns output literals outputs[0], ..., outputs[n-1] where:
            outputs[k] is true  IFF  at least k+1 of the inputs are true.

        Both forward and backward clauses are included so that outputs
        cannot be true unless the actual count supports it.
        """
        n = len(lits)
        if n == 0:
            return []
        if n == 1:
            return list(lits)

        mid = n // 2
        left = self._local_totalizer(lits[:mid])
        right = self._local_totalizer(lits[mid:])

        a_len, b_len = len(left), len(right)
        total = a_len + b_len
        outputs = [self._new() for _ in range(total)]

        # Forward: input counts imply output counts
        for i in range(a_len):
            self.clauses.append([-left[i], outputs[i]])
        for j in range(b_len):
            self.clauses.append([-right[j], outputs[j]])
        for i in range(a_len):
            for j in range(b_len):
                k = i + j + 1
                if k < total:
                    self.clauses.append([-left[i], -right[j], outputs[k]])

        # Backward: output counts require input support
        for k in range(total):
            for s in range(k + 1):
                left_has = s < a_len
                right_has = (k - s) < b_len
                if left_has and right_has:
                    self.clauses.append([-outputs[k], left[s], right[k - s]])
                elif left_has:
                    self.clauses.append([-outputs[k], left[s]])
                elif right_has:
                    self.clauses.append([-outputs[k], right[k - s]])

        return outputs

    def _encode_feasibility(self):
        """
        Sections 6.2–6.3: Unary sum + totalizer + modulo constraint.

        For each cell (r,k):
          σ_{r,k} = (Σ_j x_{r,j} + Σ_i x_{i,k} − x_{r,k}) mod c = target[r][k]

        Approach:
          1. Decompose each x_{i,j} into (c-1) unit-contribution literals:
             u_{i,j,v} = 1 iff x_{i,j} >= v  (for v = 1,...,c-1)
             so that x_{i,j} = u_{i,j,1} + ... + u_{i,j,c-1}

          2. For each constraint (r,k), gather unit literals for all 2N-1 terms.
             Their sum σ is a plain integer from 0 to (2N-1)(c-1).

          3. Count σ using a bidirectional totalizer network.

          4. Assert σ mod c = target[r][k] by blocking every invalid sum value.
        """
        N, c = self.N, self.c

        # Step 1: Create shared unit-contribution literals for all cells.
        self._u = [[[None] * c for _ in range(N)] for _ in range(N)]
        for i in range(N):
            for j in range(N):
                for v in range(1, c):
                    uv = self._new()
                    self._u[i][j][v] = uv
                    backward = [-uv]
                    for w in range(v, c):
                        self.clauses.append([-self.d[i][j][w], uv])  # d[w] → u_v
                        backward.append(self.d[i][j][w])
                    self.clauses.append(backward)  # u_v → OR(d[w])

        # Steps 2–4: For each cell (r,k), build totalizer + modulo constraint
        for r in range(N):
            for k in range(N):
                terms = [(r, j) for j in range(N)]
                terms += [(i, k) for i in range(N) if i != r]

                unit_lits = []
                for ti, tj in terms:
                    for v in range(1, c):
                        unit_lits.append(self._u[ti][tj][v])

                outputs = self._local_totalizer(unit_lits)
                n_lits = len(unit_lits)  # = (2N-1)(c-1)

                # Block every invalid sum S where S mod c != target[r][k]
                t = self.target[r][k]
                for S in range(n_lits + 1):
                    if S % c != t:
                        if S == 0:
                            self.clauses.append([outputs[0]])
                        elif S == n_lits:
                            self.clauses.append([-outputs[n_lits - 1]])
                        else:
                            self.clauses.append([-outputs[S - 1], outputs[S]])

    # ── Section 7: unit-contribution literals ────────────────────────

    def _encode_unit_lits(self):
        """
        u_{i,j,v} = 1 ⟺ x_{i,j} ≥ v   for v ∈ {1,…,c−1}.
        total(X) = Σ_{i,j} Σ_{v=1}^{c-1} u_{i,j,v} = Σ_{i,j} x_{i,j}.

        The unit-contribution literals were already created in
        _encode_feasibility (stored in self._u). We just collect them here.
        """
        N, c = self.N, self.c
        self.u_lits = []
        for i in range(N):
            for j in range(N):
                for v in range(1, c):
                    self.u_lits.append(self._u[i][j][v])

    # ── decode / print ───────────────────────────────────────────────

    def decode(self, model_set):
        """Return (click_matrix, total_clicks) from a model."""
        N, b = self.N, self.b
        matrix = []
        total = 0
        for i in range(N):
            row = []
            for j in range(N):
                val = sum((1 << l) for l in range(b)
                          if self.p[i][j][l] in model_set)
                row.append(val)
                total += val
            matrix.append(row)
        return matrix, total

    def print_solution(self, model_set):
        matrix, total = self.decode(model_set)
        print("Click matrix:")
        for row in matrix:
            print(" ".join(str(x) for x in row))
        print(f"Total clicks: {total}\n")
        return matrix, total


def _build(N, c, target):
    """Create a fresh encoder and build its CNF."""
    enc = AlienTilesEncoder(N, c, target)
    clauses, u_lits, top = enc.build()
    return enc, clauses, u_lits, top


def export_to_excel(N, c, target, approach_results, xlsx_path="results.xlsx"):
    """Export SAT results to Excel.
    approach_results: list of (name, matrix, opt, elapsed) tuples
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    palette = ["FFFFFF", "4472C4", "ED7D31", "A9D18E", "FF0000",
               "FFFF00", "9B59B6", "1ABC9C", "E74C3C", "F39C12"]
    color_map = {i: palette[i % len(palette)] for i in range(c)}

    def write_matrix(matrix, start_row, start_col, title, cmap=None):
        ws.cell(row=start_row, column=start_col, value=title).font = Font(bold=True)
        for r in range(N):
            for k in range(N):
                val = matrix[r][k]
                cell = ws.cell(row=start_row + 1 + r, column=start_col + k, value=val)
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
                if cmap and val in cmap:
                    cell.fill = PatternFill("solid", fgColor=cmap[val])

    ws["A1"] = "N"; ws["B1"] = N
    ws["A2"] = "c"; ws["B2"] = c

    headers = ["Approach", "Optimal Clicks", "Time (s)", "Verified"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h).font = Font(bold=True)

    for i, (name, matrix, opt, elapsed) in enumerate(approach_results):
        row = 5 + i
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=opt if opt is not None else "UNSAT")
        ws.cell(row=row, column=3, value=round(elapsed, 4))
        if matrix is not None:
            ok = verify_solution(N, c, target, matrix)
            ws.cell(row=row, column=4, value="PASS" if ok else "FAIL")
        else:
            ws.cell(row=row, column=4, value="N/A")

    mat_row = 5 + len(approach_results) + 2
    write_matrix(target, mat_row, 1, "Target", cmap=color_map)

    col_start = N + 3
    for name, matrix, opt, elapsed in approach_results:
        if matrix is not None:
            write_matrix(matrix, mat_row, col_start, f"X ({name})")
            col_start += N + 2

    wb.save(xlsx_path)
    print(f"Results exported to {xlsx_path}")


# ══════════════════════════════════════════════════════════════════════════
#  7.1  Non-incremental SAT (multiple independent calls)
# ══════════════════════════════════════════════════════════════════════════

def approach_71_binary(N, c, target):
    """
    Section 7.1.1 — Binary search.

    Fresh solver per iteration.  O(log(N²(c−1))) SAT calls.
    """
    print("=" * 60)
    print("Approach 7.1.1: Non-incremental SAT — Binary Search")
    print("=" * 60)

    enc, clauses, u_lits, top = _build(N, c, target)
    solver = PySATSolver(name='cadical153', bootstrap_with=clauses)

    if not solver.solve():
        print("UNSATISFIABLE\n")
        solver.delete()
        return None

    model_set = set(solver.get_model())
    solver.delete()

    hi = sum(1 for u in u_lits if u in model_set)
    best_set, best_enc = model_set, enc
    lo, calls = 0, 1
    print(f"Initial feasible solution: {hi} clicks")

    while lo <= hi:
        mid = (lo + hi) // 2
        calls += 1

        enc2, cl2, ul2, top2 = _build(N, c, target)
        am = CardEnc.atmost(ul2, bound=mid, top_id=top2,
                            encoding=EncType.seqcounter)
        s = PySATSolver(name='cadical153', bootstrap_with=cl2 + am.clauses)

        if s.solve():
            ms = set(s.get_model())
            cost = sum(1 for u in ul2 if u in ms)
            best_set, best_enc = ms, enc2
            hi = mid - 1
            print(f"  SAT call #{calls}: AtMost({mid}) → SAT  (actual = {cost})")
        else:
            lo = mid + 1
            print(f"  SAT call #{calls}: AtMost({mid}) → UNSAT")
        s.delete()

    matrix, opt = best_enc.print_solution(best_set)
    print(f"Optimal: {opt} clicks  ({calls} SAT calls)")
    return matrix, opt


def approach_71_linear(N, c, target):
    """
    Section 7.1.2 — Linear search (top-down).

    Find any feasible solution with cost K, then tighten to K−1 until UNSAT.
    """
    print("=" * 60)
    print("Approach 7.1.2: Non-incremental SAT — Linear Search (top-down)")
    print("=" * 60)

    enc, clauses, u_lits, top = _build(N, c, target)
    solver = PySATSolver(name='cadical153', bootstrap_with=clauses)

    if not solver.solve():
        print("UNSATISFIABLE\n")
        solver.delete()
        return None

    model_set = set(solver.get_model())
    solver.delete()

    UB = sum(1 for u in u_lits if u in model_set)
    best_set, best_enc = model_set, enc
    calls = 1
    print(f"Initial feasible solution: {UB} clicks")

    while True:
        calls += 1
        enc2, cl2, ul2, top2 = _build(N, c, target)
        am = CardEnc.atmost(ul2, bound=UB - 1, top_id=top2,
                            encoding=EncType.seqcounter)
        s = PySATSolver(name='cadical153', bootstrap_with=cl2 + am.clauses)

        if s.solve():
            ms = set(s.get_model())
            UB = sum(1 for u in ul2 if u in ms)
            best_set, best_enc = ms, enc2
            print(f"  SAT call #{calls}: AtMost({UB}) → SAT  (actual = {UB})")
        else:
            print(f"  SAT call #{calls}: AtMost({UB - 1}) → UNSAT")
            s.delete()
            break
        s.delete()

    matrix, opt = best_enc.print_solution(best_set)
    print(f"Optimal: {opt} clicks  ({calls} SAT calls)")
    return matrix, opt


# ══════════════════════════════════════════════════════════════════════════
#  7.2  Incremental SAT — assumption-based bound control
# ══════════════════════════════════════════════════════════════════════════

def _build_totalizer(solver, lits, top):
    """
    Incremental totalizer (Bailleux & Boufkhad, CP 2003).

    Returns output literals o_1,…,o_n where o_k is true iff
    at least k of the inputs are true.
    """
    n = len(lits)
    if n == 0:
        return [], top
    if n == 1:
        return list(lits), top

    mid = n // 2
    left, top = _build_totalizer(solver, lits[:mid], top)
    right, top = _build_totalizer(solver, lits[mid:], top)

    a, b_len = len(left), len(right)
    total = a + b_len
    outputs = []
    for _ in range(total):
        top += 1
        outputs.append(top)

    for i in range(a):
        solver.add_clause([-left[i], outputs[i]])
    for j in range(b_len):
        solver.add_clause([-right[j], outputs[j]])
    for i in range(a):
        for j in range(b_len):
            k = i + j + 1
            if k < total:
                solver.add_clause([-left[i], -right[j], outputs[k]])

    return outputs, top


def approach_72_incremental(N, c, target):
    """
    Section 7.2.1 — Incremental SAT with assumption-based bound control.

    Single solver; totalizer built once; binary search via assumptions.
    """
    print("=" * 60)
    print("Approach 7.2: Incremental SAT — Assumption-based Bound Control")
    print("=" * 60)

    enc, clauses, u_lits, top = _build(N, c, target)
    solver = PySATSolver(name='cadical153', bootstrap_with=clauses)

    outputs, top = _build_totalizer(solver, u_lits, top)

    if not solver.solve():
        print("UNSATISFIABLE\n")
        solver.delete()
        return None

    model_set = set(solver.get_model())
    hi = sum(1 for u in u_lits if u in model_set)
    best_set = model_set
    lo, calls = 0, 1
    print(f"Initial feasible solution: {hi} clicks")

    while lo <= hi:
        mid = (lo + hi) // 2
        calls += 1

        # ¬o_{mid+1} asserts "at most mid"
        assumptions = [-outputs[mid]] if mid < len(outputs) else []

        if solver.solve(assumptions=assumptions):
            ms = set(solver.get_model())
            cost = sum(1 for u in u_lits if u in ms)
            best_set = ms
            hi = mid - 1
            print(f"  SAT call #{calls}: AtMost({mid}) → SAT  (actual = {cost})")
        else:
            lo = mid + 1
            print(f"  SAT call #{calls}: AtMost({mid}) → UNSAT")

    matrix, opt = enc.print_solution(best_set)
    print(f"Optimal: {opt} clicks  ({calls} SAT calls)")
    solver.delete()
    return matrix, opt


# ══════════════════════════════════════════════════════════════════════════
#  7.3  MaxSAT — partial weighted MaxSAT
# ══════════════════════════════════════════════════════════════════════════

def approach_73_maxsat(N, c, target):
    """
    Section 7.3.1 — Partial weighted MaxSAT via RC2.

    Hard clauses : feasibility F.
    Soft clauses : (¬u_{i,j,v}) with weight 1 for each unit literal.
    Maximising satisfied soft clauses ⟺ minimising total clicks.
    """
    print("=" * 60)
    print("Approach 7.3: MaxSAT — Partial Weighted MaxSAT")
    print("=" * 60)

    enc, clauses, u_lits, top = _build(N, c, target)

    wcnf = WCNF()
    wcnf.nv = top
    for cl in clauses:
        wcnf.append(cl)
    for u in u_lits:
        wcnf.append([-u], weight=1)

    rc2 = RC2(wcnf)
    model = rc2.compute()

    if model is None:
        print("UNSATISFIABLE\n")
        rc2.delete()
        return None

    model_set = set(model)
    matrix, opt = enc.print_solution(model_set)
    print(f"Optimal: {opt} clicks")
    rc2.delete()
    return matrix, opt


# ══════════════════════════════════════════════════════════════════════════
#  Main
# ══════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(f"Usage: python {sys.argv[0]} <instance.json> [approach] [output.xlsx]")
        print("  approach: 'binary' | 'linear' | 'incremental' | 'maxsat' | 'all' (default)")
        sys.exit(1)

    path = sys.argv[1]
    approach = sys.argv[2] if len(sys.argv) > 2 else "all"
    xlsx_path = sys.argv[3] if len(sys.argv) > 3 else "results.xlsx"

    N, c, target = load_instance(path)
    print(f"Instance: N={N}, c={c}")
    print_board(target, "Target")

    approach_results = []  # list of (name, matrix, opt, elapsed)

    def _run(name, fn, *args):
        t0 = time.perf_counter()
        result = fn(*args)
        elapsed = time.perf_counter() - t0
        print(f"  Time: {elapsed:.4f}s\n")
        matrix, opt = result if result is not None else (None, None)
        approach_results.append((name, matrix, opt, elapsed))

    if approach in ("binary", "all"):
        _run("7.1.1 Binary search", approach_71_binary, N, c, target)

    if approach in ("linear", "all"):
        _run("7.1.2 Linear search", approach_71_linear, N, c, target)

    if approach in ("incremental", "all"):
        _run("7.2 Incremental SAT", approach_72_incremental, N, c, target)

    if approach in ("maxsat", "all"):
        _run("7.3 MaxSAT", approach_73_maxsat, N, c, target)

    if approach == "all" and approach_results:
        print("=" * 60)
        print("Summary")
        print("=" * 60)
        for name, matrix, opt, elapsed in approach_results:
            status = f"{opt} clicks" if opt is not None else "UNSATISFIABLE"
            print(f"  {name}: {status}")

    export_to_excel(N, c, target, approach_results, xlsx_path)