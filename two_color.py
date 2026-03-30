import json
import sys
import time

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from pycryptosat import Solver as CryptoSolver
from pysat.solvers import Solver as PySATSolver
from pysat.formula import CNF, WCNF
from pysat.card import CardEnc, EncType
from pysat.examples.rc2 import RC2

def load_instance(path: str):
    with open(path) as f:
        data = json.load(f)
    N = data["N"]
    c = data["c"]
    target = data["target"]
    return N, c, target


def print_board(board, title="Board"):
    print(f"{title}:")
    for row in board:
        print(" ".join(str(cell) for cell in row))
    print()


def print_solution(model, N, click_var):
    total = 0
    print("Click matrix:")
    for i in range(N):
        row_vals = []
        for j in range(N):
            v = click_var(i, j)
            val = 1 if model[v] else 0
            row_vals.append(val)
            total += val
        print(" ".join(str(x) for x in row_vals))
    print(f"Total clicks: {total}\n")
    return total


# ── XOR feasibility model (pycryptosat) ─────────────────────────────────

def var_id(i, j, N):
    """1-based SAT variable for click on cell (i, j)."""
    return i * N + j + 1


def build_xor_model(N, target):
    solver = CryptoSolver()

    def vv(i, j):
        return var_id(i, j, N)

    for r in range(N):
        for k in range(N):
            xor_vars = []
            # all cells in row r
            for j in range(N):
                xor_vars.append(vv(r, j))
            # all cells in column k except (r, k) which is already included
            for i in range(N):
                if i != r:
                    xor_vars.append(vv(i, k))

            rhs = bool(target[r][k])  # start state is 0
            solver.add_xor_clause(xor_vars, rhs)

    click_vars = [vv(i, j) for i in range(N) for j in range(N)]
    return solver, click_vars



def add_atmost_k(solver, variables, K):
    n = len(variables)
    if K >= n:
        return

    s = {}
    next_var = max(variables) + 1

    def new_var():
        nonlocal next_var
        v = next_var
        next_var += 1
        return v

    for i in range(n):
        for j in range(K + 1):
            s[(i, j)] = new_var()

    for i in range(n):
        # x_i → s(i, 0)
        solver.add_clause([-variables[i], s[(i, 0)]])

        if i > 0:
            # s(i-1, 0) → s(i, 0)
            solver.add_clause([-s[(i - 1, 0)], s[(i, 0)]])

        for j in range(1, K + 1):
            if i > 0:
                # s(i-1, j) → s(i, j)
                solver.add_clause([-s[(i - 1, j)], s[(i, j)]])
                # x_i ∧ s(i-1, j-1) → s(i, j)
                solver.add_clause([-variables[i], -s[(i - 1, j - 1)], s[(i, j)]])

    # forbid K+1 trues
    solver.add_clause([-s[(n - 1, K)]])


# ── XOR feasibility as CNF clauses (for pysat) ─────────────────────────

def xor_to_cnf_clauses(variables, rhs):
    n = len(variables)
    clauses = []
    for mask in range(1 << n):
        # count how many variables are negated in this assignment
        neg_count = bin(mask).count('1')
        # the clause is satisfied iff the parity of negations matches rhs
        if (neg_count % 2) != (1 if rhs else 0):
            clause = []
            for bit_idx in range(n):
                if (mask >> bit_idx) & 1:
                    clause.append(-variables[bit_idx])
                else:
                    clause.append(variables[bit_idx])
            clauses.append(clause)
    return clauses


def build_cnf_feasibility(N, target, top_var=0):
    """
    Build XOR feasibility as a plain CNF formula (for pysat).
    Returns (cnf_clauses, click_vars, top_var_used).
    """
    clauses = []

    def vv(i, j):
        return var_id(i, j, N)

    for r in range(N):
        for k in range(N):
            xor_vars = []
            for j in range(N):
                xor_vars.append(vv(r, j))
            for i in range(N):
                if i != r:
                    xor_vars.append(vv(i, k))
            rhs = bool(target[r][k])
            clauses.extend(xor_to_cnf_clauses(xor_vars, rhs))

    click_vars = [vv(i, j) for i in range(N) for j in range(N)]
    top = max(click_vars) if click_vars else 0
    if top_var > top:
        top = top_var
    return clauses, click_vars, top


# ══════════════════════════════════════════════════════════════════════════
#  7.1  Non-incremental SAT
# ══════════════════════════════════════════════════════════════════════════

def approach_71_binary_search(N, target):
    """
    Section 7.1.1 — Binary search.

    Search for the minimum K such that F ∧ AtMost({x}, K) is satisfiable.
    Each iteration builds a *fresh* solver (no clause reuse).
    Terminates in ceil(log2(N^2)) SAT calls.
    """
    print("=" * 60)
    print("Approach 7.1.1: Non-incremental SAT — Binary Search")
    print("=" * 60)

    n_vars = N * N  # total click variables

    # First check feasibility
    solver, click_vars = build_xor_model(N, target)
    sat, model = solver.solve()
    if not sat:
        print("UNSATISFIABLE — no solution exists.\n")
        return None

    lo = 0
    hi = sum(1 for v in click_vars if model[v])
    best_model = model
    calls = 0

    print(f"Initial feasible solution: {hi} clicks")

    while lo <= hi:
        mid = (lo + hi) // 2
        calls += 1

        # Fresh solver for each call
        solver, click_vars = build_xor_model(N, target)
        add_atmost_k(solver, click_vars, mid)

        sat, model = solver.solve()

        if sat:
            best_model = model
            hi = mid - 1
            cost = sum(1 for v in click_vars if model[v])
            print(f"  SAT call #{calls}: AtMost({mid}) → SAT  (actual = {cost})")
        else:
            lo = mid + 1
            print(f"  SAT call #{calls}: AtMost({mid}) → UNSAT")

    opt = sum(1 for v in click_vars if best_model[v])
    matrix = [[1 if best_model[var_id(i, j, N)] else 0 for j in range(N)] for i in range(N)]
    print(f"\nOptimal: {opt} clicks  ({calls} SAT calls)")
    print_solution(best_model, N, lambda i, j: var_id(i, j, N))
    return matrix, opt


def approach_71_linear_search(N, target):
    """
    Section 7.1.2 — Linear search (top-down).

    Find any feasible solution with cost K, then repeatedly solve with
    bound K-1 until UNSAT.  Fresh solver each iteration.
    """
    print("=" * 60)
    print("Approach 7.1.2: Non-incremental SAT — Linear Search (top-down)")
    print("=" * 60)

    solver, click_vars = build_xor_model(N, target)
    sat, model = solver.solve()
    if not sat:
        print("UNSATISFIABLE — no solution exists.\n")
        return None

    UB = sum(1 for v in click_vars if model[v])
    best_model = model
    calls = 1
    print(f"Initial feasible solution: {UB} clicks")

    while True:
        calls += 1
        solver, click_vars = build_xor_model(N, target)
        add_atmost_k(solver, click_vars, UB - 1)

        sat, model = solver.solve()
        if not sat:
            print(f"  SAT call #{calls}: AtMost({UB - 1}) → UNSAT")
            break

        best_model = model
        UB = sum(1 for v in click_vars if model[v])
        print(f"  SAT call #{calls}: AtMost({UB}) → SAT  (actual = {UB})")

    matrix = [[1 if best_model[var_id(i, j, N)] else 0 for j in range(N)] for i in range(N)]
    print(f"\nOptimal: {UB} clicks  ({calls} SAT calls)")
    print_solution(best_model, N, lambda i, j: var_id(i, j, N))
    return matrix, UB


# ══════════════════════════════════════════════════════════════════════════
#  7.2  Incremental SAT — Assumption-based bound control
# ══════════════════════════════════════════════════════════════════════════

def build_totalizer(solver, literals, top_var):
    """
    Build an incremental totalizer (Bailleux & Boufkhad, CP 2003).

    Given n input literals, produces n output literals o_1, ..., o_n where
    o_k is true iff at least k inputs are true.

    Returns (outputs, new_top_var).
    """
    n = len(literals)
    if n == 0:
        return [], top_var
    if n == 1:
        return list(literals), top_var

    # Split
    mid = n // 2
    left_out, top_var = build_totalizer(solver, literals[:mid], top_var)
    right_out, top_var = build_totalizer(solver, literals[mid:], top_var)

    # Merge: output o_k is true iff at least k of the inputs are true
    # o has indices 0..n-1 (representing "at least 1" .. "at least n")
    a = len(left_out)
    b = len(right_out)
    total = a + b

    outputs = []
    for _ in range(total):
        top_var += 1
        outputs.append(top_var)

    # Clauses:  left_out[i] ∧ right_out[j] → outputs[i+j+1]
    # (if at least i+1 from left and at least j+1 from right,
    #  then at least i+j+2 total)
    # Also propagate: left_out[i] → outputs[i] and right_out[j] → outputs[j]

    for i in range(a):
        # left_out[i] true means ≥ i+1 from left → ≥ i+1 total
        solver.add_clause([-left_out[i], outputs[i]])
    for j in range(b):
        # right_out[j] true means ≥ j+1 from right → ≥ j+1 total
        solver.add_clause([-right_out[j], outputs[j]])

    for i in range(a):
        for j in range(b):
            k = i + j + 1  # 0-indexed output = "at least k+1"
            if k < total:
                solver.add_clause([-left_out[i], -right_out[j], outputs[k]])

    return outputs, top_var


def approach_72_incremental(N, target):
    """
    Section 7.2.1 — Incremental SAT with assumption-based bound control.

    A single solver is kept alive.  The totalizer is built once.
    Binary search is performed by changing only the assumptions.
    """
    print("=" * 60)
    print("Approach 7.2: Incremental SAT — Assumption-based Bound Control")
    print("=" * 60)

    # Build feasibility CNF and add to a single pysat solver
    cnf_clauses, click_vars, top = build_cnf_feasibility(N, target)
    n = len(click_vars)  # = N^2

    solver = PySATSolver(name='cadical153', bootstrap_with=cnf_clauses)

    # Build totalizer on click_vars: outputs[k] true ⟺ at least k+1 clicks
    outputs, top = build_totalizer(solver, click_vars, top)

    # Check initial feasibility
    if not solver.solve():
        print("UNSATISFIABLE — no solution exists.\n")
        solver.delete()
        return None

    model = solver.get_model()
    model_set = set(model)

    initial_cost = sum(1 for v in click_vars if v in model_set)
    print(f"Initial feasible solution: {initial_cost} clicks")

    # Binary search using assumptions
    lo = 0
    hi = initial_cost
    best_model = model
    calls = 1

    while lo <= hi:
        mid = (lo + hi) // 2
        calls += 1

        # Assert "at most mid": assume ¬o_{mid+1}
        # outputs[mid] = "at least mid+1 are true"
        # ¬outputs[mid] = "fewer than mid+1 are true" = "at most mid"
        if mid < len(outputs):
            assumptions = [-outputs[mid]]
        else:
            assumptions = []  # no constraint needed, mid >= n

        sat = solver.solve(assumptions=assumptions)

        if sat:
            model = solver.get_model()
            model_set = set(model)
            cost = sum(1 for v in click_vars if v in model_set)
            best_model = model
            hi = mid - 1
            print(f"  SAT call #{calls}: AtMost({mid}) → SAT  (actual = {cost})")
        else:
            lo = mid + 1
            print(f"  SAT call #{calls}: AtMost({mid}) → UNSAT")

    best_set = set(best_model)
    opt = sum(1 for v in click_vars if v in best_set)
    matrix = [[1 if var_id(i, j, N) in best_set else 0 for j in range(N)] for i in range(N)]
    print(f"\nOptimal: {opt} clicks  ({calls} SAT calls)")

    # Print solution
    print("Click matrix:")
    total = 0
    for i in range(N):
        row_vals = []
        for j in range(N):
            v = var_id(i, j, N)
            val = 1 if v in best_set else 0
            row_vals.append(val)
            total += val
        print(" ".join(str(x) for x in row_vals))
    print(f"Total clicks: {total}\n")

    solver.delete()
    return matrix, opt


# ══════════════════════════════════════════════════════════════════════════
#  7.3  MaxSAT — Partial weighted MaxSAT
# ══════════════════════════════════════════════════════════════════════════

def approach_73_maxsat(N, target):
    """
    Section 7.3.1 — Partial weighted MaxSAT.

    Hard clauses: feasibility constraints F (XOR as CNF).
    Soft clauses: for each click variable x_{i,j}, add (¬x_{i,j}) with weight 1.
      Maximising satisfied soft clauses ⟺ minimising total clicks.

    Solved via RC2 (a core-guided MaxSAT solver from pysat).
    """
    print("=" * 60)
    print("Approach 7.3: MaxSAT — Partial Weighted MaxSAT")
    print("=" * 60)

    cnf_clauses, click_vars, top = build_cnf_feasibility(N, target)

    wcnf = WCNF()
    wcnf.nv = top

    # Hard clauses (weight = top_id, i.e. infinity)
    for clause in cnf_clauses:
        wcnf.append(clause)

    # Soft clauses: prefer ¬x_{i,j} (i.e. not clicking) with weight 1
    # For c=2, u_{i,j,1} = x_{i,j}, so minimising sum x_{i,j} is equivalent
    # to maximising the number of ¬x_{i,j} soft clauses satisfied.
    for v in click_vars:
        wcnf.append([-v], weight=1)

    rc2 = RC2(wcnf)
    model = rc2.compute()

    if model is None:
        print("UNSATISFIABLE — no solution exists.\n")
        rc2.delete()
        return None

    model_set = set(model)
    opt = sum(1 for v in click_vars if v in model_set)
    matrix = [[1 if var_id(i, j, N) in model_set else 0 for j in range(N)] for i in range(N)]
    print(f"Optimal: {opt} clicks")

    # Print solution
    print("Click matrix:")
    total = 0
    for i in range(N):
        row_vals = []
        for j in range(N):
            v = var_id(i, j, N)
            val = 1 if v in model_set else 0
            row_vals.append(val)
            total += val
        print(" ".join(str(x) for x in row_vals))
    print(f"Total clicks: {total}\n")

    rc2.delete()
    return matrix, opt

def get_click_from_crypto(model, click_vars):
        return {v for v in click_vars if v < len(model) and model[v]}


def _verify(N, c, target, X):
    for r in range(N):
        for k in range(N):
            sigma = sum(X[r][j] for j in range(N)) \
                  + sum(X[i][k] for i in range(N)) \
                  - X[r][k]
            if sigma % c != target[r][k]:
                return False
    return True


def export_to_excel(N, c, target, approach_results, xlsx_path="results.xlsx"):
    """Export SAT results to Excel.
    approach_results: list of (name, matrix, opt, elapsed) tuples
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    palette = ["FFFFFF", "4472C4"]  # c=2: white=0, blue=1

    def write_matrix(matrix, start_row, start_col, title, use_color=False):
        ws.cell(row=start_row, column=start_col, value=title).font = Font(bold=True)
        for r in range(N):
            for k in range(N):
                val = matrix[r][k]
                cell = ws.cell(row=start_row + 1 + r, column=start_col + k, value=val)
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
                if use_color and 0 <= val < len(palette):
                    cell.fill = PatternFill("solid", fgColor=palette[val])

    ws["A1"] = "N"; ws["B1"] = N
    ws["A2"] = "c"; ws["B2"] = c

    headers = ["Approach", "Optimal Clicks", "Time (s)", "Verified"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h).font = Font(bold=True)

    for i, (name, matrix, opt, elapsed) in enumerate(approach_results):
        row = 5 + i
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=opt if opt is not None else "UNSAT")
        ws.cell(row=row, column=3, value=round(elapsed, 4))
        if matrix is not None:
            ok = _verify(N, c, target, matrix)
            ws.cell(row=row, column=4, value="PASS" if ok else "FAIL")
        else:
            ws.cell(row=row, column=4, value="N/A")

    mat_row = 5 + len(approach_results) + 2
    write_matrix(target, mat_row, 1, "Target", use_color=True)

    col_start = N + 3
    for name, matrix, opt, elapsed in approach_results:
        if matrix is not None:
            write_matrix(matrix, mat_row, col_start, f"X ({name})")
            col_start += N + 2

    wb.save(xlsx_path)
    print(f"Results exported to {xlsx_path}")


# ══════════════════════════════════════════════════════════════════════════
#  Main
# ══════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(f"Usage: python {sys.argv[0]} <instance.json> [approach] [output.xlsx]")
        print("  approach: 'binary' | 'linear' | 'incremental' | 'maxsat' | 'all' (default)")
        sys.exit(1)

    path = sys.argv[1]
    approach = sys.argv[2] if len(sys.argv) > 2 else "all"
    xlsx_path = sys.argv[3] if len(sys.argv) > 3 else "results.xlsx"

    N, c, target = load_instance(path)
    assert c == 2, f"This solver is for c = 2, got c = {c}"

    print(f"Instance: N={N}, c={c}")
    print_board(target, "Target")

    approach_results = []  # list of (name, matrix, opt, elapsed)

    def _run(name, fn, *args):
        t0 = time.perf_counter()
        result = fn(*args)
        elapsed = time.perf_counter() - t0
        print(f"  Time: {elapsed:.4f}s\n")
        matrix, opt = result if result is not None else (None, None)
        approach_results.append((name, matrix, opt, elapsed))

    if approach in ("binary", "all"):
        _run("7.1.1 Binary search", approach_71_binary_search, N, target)

    if approach in ("linear", "all"):
        _run("7.1.2 Linear search", approach_71_linear_search, N, target)

    if approach in ("incremental", "all"):
        _run("7.2 Incremental SAT", approach_72_incremental, N, target)

    if approach in ("maxsat", "all"):
        _run("7.3 MaxSAT", approach_73_maxsat, N, target)

    if approach == "all" and approach_results:
        print("=" * 60)
        print("Summary")
        print("=" * 60)
        for name, matrix, opt, elapsed in approach_results:
            status = f"{opt} clicks" if opt is not None else "UNSATISFIABLE"
            print(f"  {name}: {status}")

    export_to_excel(N, c, target, approach_results, xlsx_path)