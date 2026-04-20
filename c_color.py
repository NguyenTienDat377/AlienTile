#!/usr/bin/env python3
"""
SAT-Based Optimisation for the Alien Tiles Problem (general c >= 2).

Implements the SAT encoding from Section 6 and the three optimisation
approaches from Section 7 of the problem specification:

  Section 6 — SAT / Pseudo-Boolean Model
    6.1  Binary encoding of click variables
    6.2  Linearised sum variables (running modular sum)
    6.3  Modulo constraint in CNF

  Section 7 — SAT-Based Optimisation Approaches
    7.1  Non-incremental SAT  (binary search + linear search top-down)
    7.2  Incremental SAT      (assumption-based bound control with totalizer)
    7.3  MaxSAT / PBO         (partial weighted MaxSAT)

Usage:
    python c_color.py <instance.json> [approach]
    approach: 'binary' | 'linear' | 'incremental' | 'maxsat' | 'all' (default)

Examples:
    python c_color.py data/4x4_c3_easy.json
    python c_color.py data/4x4_c3_easy.json binary
    python c_color.py data/4x4_c3_easy.json all
"""

import json
import math
import sys
import time
from multiprocessing import Process, Queue

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from pysat.card import CardEnc, EncType
from pysat.examples.rc2 import RC2
from pysat.formula import CNF, WCNF
from pysat.solvers import Solver as PySATSolver


# ── Helpers ──────────────────────────────────────────────────────────────

def load_instance(path: str):
    """Load an Alien Tiles instance from a JSON file."""
    with open(path) as f:
        data = json.load(f)
    return data["N"], data["c"], data["target"]


def print_board(board, title="Board"):
    """Pretty-print a 2D grid."""
    print(f"{title}:")
    for row in board:
        print("  " + " ".join(f"{cell}" for cell in row))
    print()


def verify_solution(N, c, target, X):
    """Check that click matrix X produces target T under modular arithmetic."""
    for r in range(N):
        for k in range(N):
            sigma = sum(X[r][j] for j in range(N)) \
                  + sum(X[i][k] for i in range(N)) \
                  - X[r][k]
            if sigma % c != target[r][k]:
                print(f"  MISMATCH at ({r},{k}): "
                      f"sigma={sigma}, sigma mod {c}={sigma % c}, "
                      f"target={target[r][k]}")
                return False
    return True


# ══════════════════════════════════════════════════════════════════════════
#  Section 6: SAT / Pseudo-Boolean Encoder
# ══════════════════════════════════════════════════════════════════════════

class AlienTilesEncoder:
    """
    Encodes an Alien Tiles instance (N, c, target) as a CNF formula
    using the binary encoding from Sections 6.1-6.3.
    """

    def __init__(self, N, c, target, symmetry_breaking=False):
        self.N = N
        self.c = c
        self.target = target
        self.b = math.ceil(math.log2(c)) if c > 1 else 1  # bits per click var
        self.top = 0          # next available SAT variable ID
        self.clauses = []     # list of CNF clauses
        self.p = None         # p[i][j][l]: SAT var for bit l of x_{i,j}
        self.d = None         # d[i][j][w]: SAT var for "x_{i,j} = w"
        self.u_lits = None    # unit-contribution literals for objective
        self.symmetry_breaking = symmetry_breaking

    def _new(self):
        """Allocate a fresh SAT variable."""
        self.top += 1
        return self.top

    def build(self):
        """
        Build the complete CNF:
          1. Binary click variables (Section 6.1)
          2. Value-indicator variables d[i][j][w]
          3. Feasibility constraints via running modular sum (Sections 6.2-6.3)
          4. Unit-contribution literals for objective (Section 7)

        Returns (clauses, u_lits, top).
        """
        self._encode_click_vars()
        self._encode_value_indicators()
        self._encode_feasibility()
        if self.symmetry_breaking:
            self._encode_symmetry_breaking()
        self._encode_unit_lits()
        return self.clauses, self.u_lits, self.top

    # ── Section 6.1: Binary encoding of click variables ──────────────

    def _encode_click_vars(self):
        """
        Each x_{i,j} in {0,...,c-1} is encoded using b = ceil(log2(c)) bits:
            x_{i,j} = sum_{l=0}^{b-1} 2^l * p_{i,j,l}

        If c is not a power of 2, clauses exclude invalid bit-patterns
        for values c, c+1, ..., 2^b - 1.
        """
        N, c, b = self.N, self.c, self.b

        # Allocate bit variables p[i][j][l]
        self.p = [[[self._new() for _ in range(b)]
                    for _ in range(N)]
                   for _ in range(N)]

        # Exclude invalid bit-patterns when c is not a power of 2
        if c < (1 << b):
            for i in range(N):
                for j in range(N):
                    for v in range(c, 1 << b):
                        clause = []
                        for l in range(b):
                            if (v >> l) & 1:
                                clause.append(-self.p[i][j][l])
                            else:
                                clause.append(self.p[i][j][l])
                        self.clauses.append(clause)

    # ── Value-indicator variables d[i][j][w] <=> "x_{i,j} = w" ──────

    def _encode_value_indicators(self):
        """
        For each cell (i,j) and value w in {0,...,c-1}, create a Boolean
        variable d[i][j][w] that is true iff x_{i,j} = w.

        Channelling constraints:
          Forward:  d[i][j][w] -> (each bit of p[i][j] matches w's pattern)
          Backward: (all bits match w) -> d[i][j][w]
        """
        N, c, b = self.N, self.c, self.b
        self.d = [[[None] * c for _ in range(N)] for _ in range(N)]

        for i in range(N):
            for j in range(N):
                for w in range(c):
                    dv = self._new()
                    self.d[i][j][w] = dv

                    # Forward: dv -> each bit matches w's pattern
                    backward = [dv]  # will become backward implication clause
                    for l in range(b):
                        if (w >> l) & 1:
                            self.clauses.append([-dv, self.p[i][j][l]])
                            backward.append(-self.p[i][j][l])
                        else:
                            self.clauses.append([-dv, -self.p[i][j][l]])
                            backward.append(self.p[i][j][l])

                    # Backward: all bits match w -> dv
                    self.clauses.append(backward)

    # ── Sections 6.2-6.3: Feasibility via unary sum + totalizer ──────

    def _local_totalizer(self, lits):
        """
        Build a BIDIRECTIONAL totalizer network.

        Returns output literals outputs[0], ..., outputs[n-1] where:
            outputs[k] is true  IFF  at least k+1 of the inputs are true.

        Unlike a standard (forward-only) totalizer, this includes backward
        clauses so that outputs[k] cannot be true unless the actual count
        supports it. This is required for the modulo-blocking approach.

        Forward:  actual count >= k+1  -->  outputs[k] = true
        Backward: outputs[k] = true    -->  actual count >= k+1
        """
        n = len(lits)
        if n == 0:
            return []
        if n == 1:
            return list(lits)

        mid = n // 2
        left = self._local_totalizer(lits[:mid])
        right = self._local_totalizer(lits[mid:])

        a_len, b_len = len(left), len(right)
        total = a_len + b_len
        outputs = [self._new() for _ in range(total)]

        # ── Forward: input counts imply output counts ──
        for i in range(a_len):
            self.clauses.append([-left[i], outputs[i]])
        for j in range(b_len):
            self.clauses.append([-right[j], outputs[j]])
        for i in range(a_len):
            for j in range(b_len):
                k = i + j + 1
                if k < total:
                    self.clauses.append([-left[i], -right[j], outputs[k]])

        # ── Backward: output counts require input support ──
        # For each output level k and each "cut point" s, assert:
        #   outputs[k] --> left[s] OR right[k-s]
        # This means: if combined count >= k+1, then for every way
        # to split k+1 into (s+1) from left and (k-s) from right,
        # at least one side must have enough.
        for k in range(total):
            for s in range(k + 1):
                left_has = s < a_len
                right_has = (k - s) < b_len
                if left_has and right_has:
                    self.clauses.append([-outputs[k], left[s], right[k - s]])
                elif left_has:
                    self.clauses.append([-outputs[k], left[s]])
                elif right_has:
                    self.clauses.append([-outputs[k], right[k - s]])

        return outputs

    def _encode_feasibility(self):
        """
        Sections 6.2-6.3: Unary sum + totalizer + modulo constraint.

        For each cell (r,k), the effect equation is:
            sigma_{r,k} = (sum_j x_{r,j} + sum_i x_{i,k} - x_{r,k}) mod c = target[r][k]

        Approach (follows the PDF Sections 6.2-6.3 directly):
          1. Decompose each click variable x_{i,j} into (c-1) unit-contribution
             literals:  u_{i,j,v} = 1 iff x_{i,j} >= v  (for v = 1,...,c-1)
             so that x_{i,j} = u_{i,j,1} + u_{i,j,2} + ... + u_{i,j,c-1}

          2. For each constraint (r,k), gather the unit literals for all 2N-1
             terms. Their sum sigma = sum of all unit lits (a plain integer
             from 0 to (2N-1)(c-1)).

          3. Count sigma using a totalizer network.

          4. Assert sigma mod c = target[r][k] by blocking every invalid sum
             value S where S mod c != target[r][k].
        """
        N, c = self.N, self.c

        # Step 1: Create shared unit-contribution literals for all cells.
        # self._u[i][j][v] = "x_{i,j} >= v"  for v in {1, ..., c-1}
        # Channelling:  u_v <=> OR(d[i][j][w] for w in {v,...,c-1})
        self._u = [[[None] * c for _ in range(N)] for _ in range(N)]
        for i in range(N):
            for j in range(N):
                for v in range(1, c):
                    uv = self._new()
                    self._u[i][j][v] = uv
                    backward = [-uv]
                    for w in range(v, c):
                        self.clauses.append([-self.d[i][j][w], uv])  # d[w] -> u_v
                        backward.append(self.d[i][j][w])
                    self.clauses.append(backward)  # u_v -> OR(d[w])

        # Steps 2-4: For each cell (r,k), build totalizer + modulo constraint
        for r in range(N):
            for k in range(N):
                # 2N-1 terms: full row r, then column k excluding (r,k)
                terms = [(r, j) for j in range(N)]
                terms += [(i, k) for i in range(N) if i != r]

                # Gather unit-contribution literals for these terms
                unit_lits = []
                for ti, tj in terms:
                    for v in range(1, c):
                        unit_lits.append(self._u[ti][tj][v])

                # Build totalizer: outputs[k] = "at least k+1 unit lits are true"
                outputs = self._local_totalizer(unit_lits)
                n_lits = len(unit_lits)  # = (2N-1)(c-1)

                # Block every invalid sum S where S mod c != target[r][k].
                # "sum != S" is a single 1- or 2-literal clause:
                #   S = 0:      [outputs[0]]               (force sum >= 1)
                #   0 < S < n:  [-outputs[S-1], outputs[S]] (force sum < S or sum > S)
                #   S = n:      [-outputs[n-1]]             (force sum <= n-1)
                t = self.target[r][k]
                for S in range(n_lits + 1):
                    if S % c != t:
                        if S == 0:
                            self.clauses.append([outputs[0]])
                        elif S == n_lits:
                            self.clauses.append([-outputs[n_lits - 1]])
                        else:
                            self.clauses.append([-outputs[S - 1], outputs[S]])

    # ── Section 5.4: Symmetry-breaking constraints (optional) ────────

    def _encode_symmetry_breaking(self):
        """
        Section 5.4: Optional symmetry-breaking constraints.

        Breaks the N! × N! × 2 symmetry group via three constraints:
          (8) Row-sum ordering:    sum_j x_{1,j} ≤ ... ≤ sum_j x_{N,j}
          (9) Column-sum ordering: sum_i x_{i,1} ≤ ... ≤ sum_i x_{i,N}
          (10) Diagonal reflection: x_{1,2} ≤ x_{2,1}

        For (8)/(9): build a totalizer over each row/column's unit literals,
        then add output[i][k] → output[i+1][k] for every threshold k.

        For (10): u[i][j][v] = (x_{i,j} ≥ v), so x_{1,2} ≤ x_{2,1} is
        exactly u[0][1][v] → u[1][0][v] for each v in {1,...,c-1}.
        """
        N, c = self.N, self.c
        if N < 2:
            return

        # (8) Row-sum ordering
        row_outputs = []
        for i in range(N):
            row_lits = [self._u[i][j][v]
                        for j in range(N) for v in range(1, c)]
            row_outputs.append(self._local_totalizer(row_lits))

        for i in range(N - 1):
            n = min(len(row_outputs[i]), len(row_outputs[i + 1]))
            for k in range(n):
                # row_sum[i] >= k+1  =>  row_sum[i+1] >= k+1
                self.clauses.append([-row_outputs[i][k], row_outputs[i + 1][k]])

        # (9) Column-sum ordering
        col_outputs = []
        for j in range(N):
            col_lits = [self._u[i][j][v]
                        for i in range(N) for v in range(1, c)]
            col_outputs.append(self._local_totalizer(col_lits))

        for j in range(N - 1):
            n = min(len(col_outputs[j]), len(col_outputs[j + 1]))
            for k in range(n):
                # col_sum[j] >= k+1  =>  col_sum[j+1] >= k+1
                self.clauses.append([-col_outputs[j][k], col_outputs[j + 1][k]])

        # (10) Diagonal reflection: x_{1,2} <= x_{2,1}  (0-indexed: [0][1] vs [1][0])
        for v in range(1, c):
            self.clauses.append([-self._u[0][1][v], self._u[1][0][v]])

    # ── Section 7: Unit-contribution literals for objective ──────────

    def _encode_unit_lits(self):
        """
        Unary expansion of the objective (Section 7).

        For each cell (i,j) and each v in {1,...,c-1}, define:
            u_{i,j,v} = 1  iff  x_{i,j} >= v

        Then total(X) = sum_{i,j} sum_{v=1}^{c-1} u_{i,j,v} = sum_{i,j} x_{i,j}.

        The unit-contribution literals were already created and channelled
        in _encode_feasibility (stored in self._u). We just collect them here.
        """
        N, c = self.N, self.c
        self.u_lits = []
        for i in range(N):
            for j in range(N):
                for v in range(1, c):
                    self.u_lits.append(self._u[i][j][v])

    # ── Decode solution from SAT model ───────────────────────────────

    def decode(self, model_set):
        """
        Extract (click_matrix, total_clicks) from a SAT model.

        Reads the binary encoding p[i][j][l] to reconstruct each x_{i,j}.
        """
        N, b = self.N, self.b
        matrix = []
        total = 0
        for i in range(N):
            row = []
            for j in range(N):
                val = sum((1 << l) for l in range(b)
                          if self.p[i][j][l] in model_set)
                row.append(val)
                total += val
            matrix.append(row)
        return matrix, total

    def print_solution(self, model_set):
        """Decode and print a solution from a SAT model."""
        matrix, total = self.decode(model_set)
        print("  Click matrix:")
        for row in matrix:
            print("    " + " ".join(f"{x}" for x in row))
        print(f"  Total clicks: {total}")
        return matrix, total


def _build(N, c, target, symmetry_breaking=False):
    """Create a fresh encoder and build its CNF. Returns (enc, clauses, u_lits, top)."""
    enc = AlienTilesEncoder(N, c, target, symmetry_breaking=symmetry_breaking)
    clauses, u_lits, top = enc.build()
    return enc, clauses, u_lits, top


def export_to_excel(instance_name, approach_results, xlsx_path="sat.xlsx"):
    """Export SAT results to Excel, one sheet per approach (solver + method).

    approach_results: list of (sheet_name, matrix, opt, elapsed, num_vars, num_clauses) tuples
    Appends a new row to each sheet if the workbook already exists.
    """
    import os
    if os.path.exists(xlsx_path):
        wb = openpyxl.load_workbook(xlsx_path)
    else:
        wb = openpyxl.Workbook()
        # Remove the default empty sheet created by openpyxl
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    headers = ["Instance", "Variables", "Clauses", "Runtime", "Optimal Click"]

    for sheet_name, matrix, opt, elapsed, num_vars, num_clauses in approach_results:
        # Get or create the sheet for this approach
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            next_row = ws.max_row + 1
        else:
            ws = wb.create_sheet(title=sheet_name)
            # Write header row
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True)
            next_row = 2

        timed_out = elapsed == "TIMEOUT"

        ws.cell(row=next_row, column=1, value=instance_name)
        ws.cell(row=next_row, column=2, value="Timeout" if timed_out else num_vars)
        ws.cell(row=next_row, column=3, value="Timeout" if timed_out else num_clauses)
        ws.cell(row=next_row, column=4, value="Timeout" if timed_out else round(float(elapsed), 4))
        ws.cell(row=next_row, column=5, value="Timeout" if timed_out else (opt if opt is not None else "UNSAT"))

    wb.save(xlsx_path)
    print(f"Results exported to {xlsx_path}")


# ══════════════════════════════════════════════════════════════════════════
#  Section 7.1: Non-incremental SAT (multiple independent calls)
# ══════════════════════════════════════════════════════════════════════════

def approach_71_binary(N, c, target, symmetry_breaking=False):
    """
    Section 7.1.1 — Binary search.

    Search for the minimum K such that F AND AtMost({u}, K) is satisfiable.
    A fresh solver is created for every iteration.
    O(ceil(log2(N^2(c-1)))) SAT calls.

    Algorithm:
      1. lo <- 0;  h <- N^2(c-1);  X* <- None
      2. while lo <= h:
           m <- floor((lo + h) / 2)
           Build fresh CNF: F AND AtMost({u}, m)
           if SAT:  X* <- Extract(); h <- m - 1
           else:    lo <- m + 1
      3. return X*
    """
    print("=" * 60)
    print("Approach 7.1.1: Non-incremental SAT -- Binary Search")
    print("=" * 60)
    t_start = time.perf_counter()

    # Check feasibility first (no cardinality bound)
<<<<<<< HEAD
    enc, clauses, u_lits, top = _build(N, c, target)
=======
    enc, clauses, u_lits, top = _build(N, c, target, symmetry_breaking)
>>>>>>> 9116161 (Add CPLEX-CP)
    solver = PySATSolver(name='cadical195', bootstrap_with=clauses)

    if not solver.solve():
        print("  Result: UNSATISFIABLE -- no solution exists.\n")
        solver.delete()
        return None
    model_set = set(solver.get_model())
    solver.delete()

    # Initial upper bound from first feasible solution
    hi = sum(1 for u in u_lits if u in model_set)
    best_set, best_enc = model_set, enc
    lo, calls = 0, 1
    print(f"  Initial feasible solution: {hi} clicks")

    # Binary search
    while lo <= hi:
        mid = (lo + hi) // 2
        calls += 1

        enc2, cl2, ul2, top2 = _build(N, c, target, symmetry_breaking)
        am = CardEnc.atmost(ul2, bound=mid, top_id=top2,
                            encoding=EncType.seqcounter)
        s = PySATSolver(name='cadical195', bootstrap_with=cl2 + am.clauses)

        if s.solve():
            ms = set(s.get_model())
            cost = sum(1 for u in ul2 if u in ms)
            best_set, best_enc = ms, enc2
            hi = mid - 1
            print(f"    SAT call #{calls}: AtMost({mid}) -> SAT  (actual = {cost})")
        else:
            lo = mid + 1
            print(f"    SAT call #{calls}: AtMost({mid}) -> UNSAT")
        s.delete()

    matrix, opt = best_enc.print_solution(best_set)
    elapsed = time.perf_counter() - t_start
    print(f"  Optimal: {opt} clicks  ({calls} SAT calls, {elapsed:.4f}s)")

    return matrix, opt, enc.top, len(enc.clauses)


def approach_71_linear(N, c, target, symmetry_breaking=False):
    """
    Section 7.1.2 — Linear search (top-down).

    Find any feasible solution X, record its total K, then solve with
    bound K-1, repeating until unsatisfiable.
    Needs K_first - K* + 1 calls in the worst case.

    Algorithm:
      1. Solve F (no bound). If UNSAT, stop.
      2. K <- total(X)
      3. loop:
           Build fresh CNF: F AND AtMost({u}, K-1)
           if SAT:  X* <- Extract(); K <- total(X*)
           else:    break
      4. return X*
    """
    print("=" * 60)
    print("Approach 7.1.2: Non-incremental SAT -- Linear Search (top-down)")
    print("=" * 60)
    t_start = time.perf_counter()

    # Check feasibility
<<<<<<< HEAD
    enc, clauses, u_lits, top = _build(N, c, target)
=======
    enc, clauses, u_lits, top = _build(N, c, target, symmetry_breaking)
>>>>>>> 9116161 (Add CPLEX-CP)
    solver = PySATSolver(name='cadical195', bootstrap_with=clauses)

    if not solver.solve():
        print("  Result: UNSATISFIABLE -- no solution exists.\n")
        solver.delete()
        return None
    model_set = set(solver.get_model())
    solver.delete()

    UB = sum(1 for u in u_lits if u in model_set)
    best_set, best_enc = model_set, enc
    calls = 1
    print(f"  Initial feasible solution: {UB} clicks")

    # Linear top-down search
    while True:
        calls += 1
        enc2, cl2, ul2, top2 = _build(N, c, target, symmetry_breaking)
        am = CardEnc.atmost(ul2, bound=UB - 1, top_id=top2,
                            encoding=EncType.seqcounter)
        s = PySATSolver(name='cadical153', bootstrap_with=cl2 + am.clauses)

        if s.solve():
            ms = set(s.get_model())
            UB = sum(1 for u in ul2 if u in ms)
            best_set, best_enc = ms, enc2
            print(f"    SAT call #{calls}: AtMost({UB}) -> SAT  (actual = {UB})")
        else:
            print(f"    SAT call #{calls}: AtMost({UB - 1}) -> UNSAT")
            s.delete()
            break
        s.delete()

    matrix, opt = best_enc.print_solution(best_set)
    elapsed = time.perf_counter() - t_start
    print(f"  Optimal: {opt} clicks  ({calls} SAT calls, {elapsed:.4f}s)")

    ok = verify_solution(N, c, target, matrix)
    print(f"  Verification: {'PASSED' if ok else 'FAILED'}")
    print()
    return matrix, opt, enc.top, len(enc.clauses)


# ══════════════════════════════════════════════════════════════════════════
#  Section 7.2: Incremental SAT — assumption-based bound control
# ══════════════════════════════════════════════════════════════════════════

def _build_totalizer(solver, lits, top):
    """
    Build an incremental totalizer network (Bailleux & Boufkhad, CP 2003).

    The totalizer produces output literals o_1, ..., o_n where o_k is true
    if and only if at least k of the input literals are true.

    To assert "at most K":  assume -o_{K+1}  (0-indexed: -outputs[K])
    The totalizer is built once; bound changes only modify assumptions.

    Parameters:
        solver: PySAT solver instance (clauses are added directly)
        lits:   list of input literals
        top:    current highest SAT variable ID

    Returns:
        (outputs, top) where outputs[k] means "at least k+1 inputs are true"
    """
    n = len(lits)
    if n == 0:
        return [], top
    if n == 1:
        return list(lits), top

    mid = n // 2
    left, top = _build_totalizer(solver, lits[:mid], top)
    right, top = _build_totalizer(solver, lits[mid:], top)

    a_len, b_len = len(left), len(right)
    total = a_len + b_len
    outputs = []
    for _ in range(total):
        top += 1
        outputs.append(top)

    # Merge network clauses:
    # If left[i] is true (>= i+1 from left) then outputs[i] is true
    for i in range(a_len):
        solver.add_clause([-left[i], outputs[i]])

    # If right[j] is true (>= j+1 from right) then outputs[j] is true
    for j in range(b_len):
        solver.add_clause([-right[j], outputs[j]])

    # If left[i] AND right[j] then outputs[i+j+1] is true
    for i in range(a_len):
        for j in range(b_len):
            k = i + j + 1
            if k < total:
                solver.add_clause([-left[i], -right[j], outputs[k]])

    return outputs, top


def approach_72_incremental(N, c, target, symmetry_breaking=False):
    """
    Section 7.2.1 — Incremental SAT with assumption-based bound control.

    A single solver instance is kept alive across iterations.
    The cardinality encoding (totalizer) is built once.
    Binary search on the bound K using assumptions on totalizer outputs:
      - To assert "at most K": assume NOT o_{K+1}

    Key advantages over non-incremental (Section 7.1):
      - Totalizer built ONCE (no re-encoding per call)
      - Learned clauses retained across iterations (2-10x faster)
      - Bound changes only modify assumptions, not the clause database

    Algorithm:
      1. solver <- IncrementalSolver(F)
      2. Build totalizer on {u_{i,j,v}} -> outputs o_1,...,o_n
      3. Add totalizer clauses to solver
      4. lo <- 0; h <- n; X* <- None
      5. while lo <= h:
           m <- floor((lo + h) / 2)
           if solver.Solve(assumptions = {-o_{m+1}}) then
             X* <- Extract(); h <- m - 1
           else  lo <- m + 1
      6. return X*
    """
    print("=" * 60)
    print("Approach 7.2: Incremental SAT -- Assumption-based Bound Control")
    print("=" * 60)
    t_start = time.perf_counter()

    # Build feasibility encoding + unit literals
    enc, clauses, u_lits, top = _build(N, c, target, symmetry_breaking)

    # Create a single persistent solver
    solver = PySATSolver(name='cadical195', bootstrap_with=clauses)

    # Build totalizer network ONCE
    outputs, top = _build_totalizer(solver, u_lits, top)

    # Check feasibility (no bound)
    if not solver.solve():
        print("  Result: UNSATISFIABLE -- no solution exists.\n")
        solver.delete()
        return None

    model_set = set(solver.get_model())
    hi = sum(1 for u in u_lits if u in model_set)
    best_set = model_set
    lo, calls = 0, 1
    print(f"  Initial feasible solution: {hi} clicks")

    # Binary search using assumptions
    while lo <= hi:
        mid = (lo + hi) // 2
        calls += 1

        # Assume "at most mid": negate o_{mid+1} (0-indexed: outputs[mid])
        assumptions = [-outputs[mid]] if mid < len(outputs) else []

        if solver.solve(assumptions=assumptions):
            ms = set(solver.get_model())
            cost = sum(1 for u in u_lits if u in ms)
            best_set = ms
            hi = mid - 1
            print(f"    SAT call #{calls}: AtMost({mid}) -> SAT  (actual = {cost})")
        else:
            lo = mid + 1
            print(f"    SAT call #{calls}: AtMost({mid}) -> UNSAT")

    matrix, opt = enc.print_solution(best_set)
    elapsed = time.perf_counter() - t_start
    print(f"  Optimal: {opt} clicks  ({calls} SAT calls, {elapsed:.4f}s)")

    ok = verify_solution(N, c, target, matrix)
    print(f"  Verification: {'PASSED' if ok else 'FAILED'}")

    solver.delete()
    print()
    return matrix, opt, enc.top, len(enc.clauses)


# ══════════════════════════════════════════════════════════════════════════
#  Section 7.3: MaxSAT — partial weighted MaxSAT
# ══════════════════════════════════════════════════════════════════════════

def approach_73_maxsat(N, c, target, symmetry_breaking=False):
    """
    Section 7.3.1 — Partial weighted MaxSAT via RC2.

    The formula is split into:
      - Hard clauses: all feasibility constraints F (exactly-one,
        running-sum, target assertion, value exclusion).
      - Soft clauses: for each unit literal u_{i,j,v}, add the unit
        clause (NOT u_{i,j,v}) with weight 1.

    Minimising total clicks is equivalent to maximising the number of
    satisfied soft clauses (maximising the NOT-u literals that are true,
    i.e., minimising the u literals that are true).

    The solver handles optimisation internally using core-guided
    algorithms (e.g., OLL, MSU3).
    """
    print("=" * 60)
    print("Approach 7.3: MaxSAT -- Partial Weighted MaxSAT")
    print("=" * 60)
    t_start = time.perf_counter()

    # Build feasibility encoding + unit literals
    enc, clauses, u_lits, top = _build(N, c, target, symmetry_breaking)

    # Construct WCNF (weighted CNF)
    wcnf = WCNF()
    wcnf.nv = top

    # Hard clauses: feasibility (infinite weight, added without explicit weight)
    for cl in clauses:
        wcnf.append(cl)

    # Soft clauses: (NOT u_{i,j,v}) with weight 1
    # Satisfying these means the u-literal is false, i.e., x_{i,j} < v
    for u in u_lits:
        wcnf.append([-u], weight=1)

    # Solve with RC2 (core-guided MaxSAT solver)
    rc2 = RC2(wcnf)
    model = rc2.compute()

    if model is None:
        print("  Result: UNSATISFIABLE -- no solution exists.\n")
        rc2.delete()
        return None

    model_set = set(model)
    matrix, opt = enc.print_solution(model_set)
    elapsed = time.perf_counter() - t_start
    print(f"  Optimal: {opt} clicks  ({elapsed:.4f}s)")

    ok = verify_solution(N, c, target, matrix)
    print(f"  Verification: {'PASSED' if ok else 'FAILED'}")

    rc2.delete()
    print()
    return matrix, opt, enc.top, len(enc.clauses)


# ══════════════════════════════════════════════════════════════════════════
#  Multiprocessing worker (must be module-level for pickling on macOS)
# ══════════════════════════════════════════════════════════════════════════

def _worker(queue, fn, args):
    """Run fn(*args) in a child process and put the result in queue."""
    try:
        result = fn(*args)
        queue.put(result)
    except Exception as exc:
        queue.put(exc)


# ══════════════════════════════════════════════════════════════════════════
#  Main
# ══════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(f"Usage: python {sys.argv[0]} <instance.json> [approach] [sym] [output.xlsx]")
        print("  approach: 'binary' | 'linear' | 'incremental' | 'maxsat' | 'all' (default)")
        print("  sym:      add 'sym' to enable symmetry-breaking constraints (Section 5.4)")
        sys.exit(1)

    path = sys.argv[1]
<<<<<<< HEAD
    approach = sys.argv[2] if len(sys.argv) > 2 else "all"
    xlsx_path = sys.argv[3] if len(sys.argv) > 3 else "sat.xlsx"
=======
    extra = sys.argv[2:]
    use_sym = "sym" in extra
    approach = next((a for a in extra if a in ("binary", "linear", "incremental", "maxsat", "all")), "all")
    xlsx_path = next((a for a in extra if a.endswith(".xlsx")), "sat.xlsx")
>>>>>>> 9116161 (Add CPLEX-CP)

    import os
    instance_name = os.path.splitext(os.path.basename(path))[0]

    N, c, target = load_instance(path)
    print(f"Instance: N={N}, c={c}")
    print(f"Symmetry breaking: {'ON' if use_sym else 'OFF'}")
    print_board(target, "Target")

    # ── Timeout configuration ──────────────────────────────────────────
    TIMEOUT_SECONDS = 300  # seconds; change this value to adjust the limit

    # Each entry: (sheet_name, matrix, opt, elapsed, num_vars, num_clauses)
    approach_results = []

    def _run(sheet_name, fn, *args):
        queue = Queue()
        p = Process(target=_worker, args=(queue, fn, args))
        t0 = time.perf_counter()
        p.start()
        p.join(timeout=TIMEOUT_SECONDS)

        if p.is_alive():
            p.kill()
            p.join()
            elapsed = time.perf_counter() - t0
            print(f"  [{sheet_name}] TIMEOUT after {elapsed:.1f}s — skipping.")
            approach_results.append((sheet_name, None, None, "TIMEOUT", None, None))
            return

        elapsed = time.perf_counter() - t0
        result = queue.get() if not queue.empty() else None
        if isinstance(result, Exception):
            print(f"  [{sheet_name}] ERROR: {result}")
            approach_results.append((sheet_name, None, None, "ERROR", None, None))
            return

        if result is not None:
            matrix, opt, num_vars, num_clauses = result
        else:
            matrix, opt, num_vars, num_clauses = None, None, None, None
        approach_results.append((sheet_name, matrix, opt, elapsed, num_vars, num_clauses))

    if approach in ("binary", "all"):
<<<<<<< HEAD
        _run("Binary Search", approach_71_binary, N, c, target)

    if approach in ("linear", "all"):
        _run("Linear Search", approach_71_linear, N, c, target)

    if approach in ("incremental", "all"):
        _run("Incremental", approach_72_incremental, N, c, target)

    if approach in ("maxsat", "all"):
        _run("MaxSAT", approach_73_maxsat, N, c, target)
=======
        _run("Binary Search", approach_71_binary, N, c, target, use_sym)

    if approach in ("linear", "all"):
        _run("Linear Search", approach_71_linear, N, c, target, use_sym)

    if approach in ("incremental", "all"):
        _run("Incremental", approach_72_incremental, N, c, target, use_sym)

    if approach in ("maxsat", "all"):
        _run("MaxSAT", approach_73_maxsat, N, c, target, use_sym)
>>>>>>> 9116161 (Add CPLEX-CP)

    # Summary
    if approach == "all" and approach_results:
        print("=" * 60)
        print("Summary")
        print("=" * 60)
        for sheet_name, matrix, opt, elapsed, _, _ in approach_results:
            status = f"{opt} clicks" if opt is not None else "UNSATISFIABLE"
            print(f"  {sheet_name}: {status}")
        print()

    export_to_excel(instance_name, approach_results, xlsx_path)