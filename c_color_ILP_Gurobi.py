import gurobipy as gp
from gurobipy import GRB

import json
import math
import sys
import time
import openpyxl
from openpyxl.styles import Font
from multiprocessing import Process, Queue


def export_to_excel(instance_name, sheet_name, result, elapsed, num_vars, num_constrs, path="sat.xlsx"):
    """Export ILP results to Excel, one sheet per variant (Gurobi + variant name).

    Appends a new row to the sheet if the workbook already exists.
    """
    import os
    if os.path.exists(path):
        wb = openpyxl.load_workbook(path)
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    headers = ["Instance", "Variables", "Clauses", "Runtime", "Optimal Click"]

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        next_row = ws.max_row + 1
    else:
        ws = wb.create_sheet(title=sheet_name)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)
        next_row = 2

    timed_out = elapsed == "TIMEOUT"
    ws.cell(row=next_row, column=1, value=instance_name)
    ws.cell(row=next_row, column=2, value="Timeout" if timed_out else num_vars)
    ws.cell(row=next_row, column=3, value="Timeout" if timed_out else num_constrs)
    ws.cell(row=next_row, column=4, value="Timeout" if timed_out else round(float(elapsed), 4))
    ws.cell(row=next_row, column=5, value="Timeout" if timed_out else (result["total_clicks"] if result is not None else "UNSAT"))

    wb.save(path)
    print(f"Results exported to {path}")


def load_instance(path: str):
    with open(path) as f:
        data = json.load(f)
    return data["N"], data["c"], data["target"]


def print_board(board, title="Board"):
    print(f"{title}:")
    for row in board:
        print("  " + " ".join(f"{cell}" for cell in row))
    print()


def verify_solution(N, c, target, X):
    for r in range(N):
        for k in range(N):
            sigma = sum(X[r][j] for j in range(N)) \
                  + sum(X[i][k] for i in range(N)) \
                  - X[r][k]
            if sigma % c != target[r][k]:
                print(f"  MISMATCH at ({r},{k}): "
                      f"sigma={sigma}, sigma mod {c}={sigma % c}, "
                      f"target={target[r][k]}")
                return False
    return True

class AlienTilesILP:
    def __init__(self, N, c, target=None):
        self.N = N
        self.c = c
        self.target = target
        self.x_ij = [[None] * N for _ in range(N)]
        self.q_rk = [[None] * N for _ in range(N)]
        self.t_rk = [[None] * N for _ in range(N)]
        self.mdl = gp.Model("AlienTilesILP")
        self.mdl.setParam("OutputFlag", 0)

    def _encode_variables(self) -> None:
        upper_bound = math.floor((2 * self.N - 1) * (self.c - 1) / self.c)
        for i in range(self.N):
            for j in range(self.N):
                self.x_ij[i][j] = self.mdl.addVar(
                    lb=0, ub=self.c - 1, vtype=GRB.INTEGER, name=f"x_{i}_{j}"
                )
        for r in range(self.N):
            for k in range(self.N):
                self.q_rk[r][k] = self.mdl.addVar(
                    lb=0, ub=upper_bound, vtype=GRB.INTEGER, name=f"q_{r}_{k}"
                )
        self.mdl.update()

    def _encode_constraints(self, target_vars=None) -> None:
        t = target_vars if target_vars else self.target
        for r in range(self.N):
            for k in range(self.N):
                sigma = (gp.quicksum(self.x_ij[r][j] for j in range(self.N))
                         + gp.quicksum(self.x_ij[i][k] for i in range(self.N))
                         - self.x_ij[r][k])
                self.mdl.addConstr(
                    sigma - self.c * self.q_rk[r][k] == t[r][k],
                    name=f"mod_eq_{r}_{k}"
                )

    def _encode_symmetry_breaking(self) -> None:
        N = self.N
        for i in range(N - 1):
            self.mdl.addConstr(
                gp.quicksum(self.x_ij[i][j] for j in range(N))
                <= gp.quicksum(self.x_ij[i + 1][j] for j in range(N)),
                name=f"row_sym_{i}"
            )
        for j in range(N - 1):
            self.mdl.addConstr(
                gp.quicksum(self.x_ij[i][j] for i in range(N))
                <= gp.quicksum(self.x_ij[i][j + 1] for i in range(N)),
                name=f"col_sym_{j}"
            )
        self.mdl.addConstr(self.x_ij[0][1] <= self.x_ij[1][0], name="diag_sym")

    def build_variant1(self, symmetry_breaking=False):
        self._encode_variables()
        self._encode_constraints()
        if symmetry_breaking:
            self._encode_symmetry_breaking()

    def build_variant2(self, symmetry_breaking=False):
        self._encode_variables()
        self._encode_constraints()
        if symmetry_breaking:
            self._encode_symmetry_breaking()
        self.mdl.setObjective(
            gp.quicksum(self.x_ij[i][j] for i in range(self.N) for j in range(self.N)),
            GRB.MINIMIZE
        )

    def build_variant3(self, symmetry_breaking=False):
        self._encode_variables()
        for r in range(self.N):
            for k in range(self.N):
                self.t_rk[r][k] = self.mdl.addVar(
                    lb=0, ub=self.c - 1, vtype=GRB.INTEGER, name=f"t_{r}_{k}"
                )
        self.mdl.update()
        self._encode_constraints(target_vars=self.t_rk)
        self.mdl.addConstr(
            gp.quicksum(self.t_rk[r][k]
                        for r in range(self.N) for k in range(self.N)) >= 1,
            name="non_trivial_target"
        )
        if symmetry_breaking:
            self._encode_symmetry_breaking()
        self.mdl.setObjective(
            gp.quicksum(self.x_ij[i][j] for i in range(self.N) for j in range(self.N)),
            GRB.MAXIMIZE
        )

    def solve(self):
        self.mdl.optimize()
        if self.mdl.Status != GRB.OPTIMAL:
            return None
        X = [[int(self.x_ij[i][j].X) for j in range(self.N)] for i in range(self.N)]
        T = None
        if self.t_rk[0][0] is not None:
            T = [[int(self.t_rk[r][k].X) for k in range(self.N)] for r in range(self.N)]
        total = sum(X[i][j] for i in range(self.N) for j in range(self.N))
        return {"X": X, "T": T, "total_clicks": total}


def _worker(queue, N, c, target, variant):
    """Build and solve in a child process; put (result, num_vars, num_constrs) in queue."""
    try:
        ilp = AlienTilesILP(N, c, target)
        if variant == 1:
            ilp.build_variant1()
        elif variant == 2:
            ilp.build_variant2()
        elif variant == 3:
            ilp.build_variant3()
        ilp.mdl.update()
        num_vars = ilp.mdl.NumVars
        num_constrs = ilp.mdl.NumConstrs
        result = ilp.solve()
        queue.put((result, num_vars, num_constrs))
    except Exception as exc:
        queue.put(exc)


if __name__ == "__main__":
    import os

    if len(sys.argv) < 2:
        print(f"Usage: python {sys.argv[0]} <instance.json> [variant] [output.xlsx]")
        print("  variant: 1 (feasibility) | 2 (min clicks) | 3 (hardest puzzle) | all (default)")
        sys.exit(1)

    path = sys.argv[1]
    variant_arg = sys.argv[2] if len(sys.argv) > 2 else "all"
    xlsx_path = sys.argv[3] if len(sys.argv) > 3 else "sat.xlsx"

    TIMEOUT_SECONDS = 300  # seconds; change this value to adjust the limit

    instance_name = os.path.splitext(os.path.basename(path))[0]

    N, c, target = load_instance(path)
    print(f"Instance: N={N}, c={c}")
    print_board(target, "Target")

    variants = [1, 2, 3] if variant_arg == "all" else [int(variant_arg)]

    for variant in variants:
        sheet_name = f"Gurobi Variant {variant}"
        print(f"--- Variant {variant} ---")

        queue = Queue()
        p = Process(target=_worker, args=(queue, N, c, target, variant))
        t0 = time.time()
        p.start()
        p.join(timeout=TIMEOUT_SECONDS)

        if p.is_alive():
            p.kill()
            p.join()
            elapsed = time.time() - t0
            print(f"  TIMEOUT after {elapsed:.1f}s — skipping.")
            export_to_excel(instance_name, sheet_name, None, "TIMEOUT", None, None, path=xlsx_path)
            continue

        elapsed = time.time() - t0
        payload = queue.get() if not queue.empty() else None

        if isinstance(payload, Exception):
            print(f"  ERROR: {payload}")
            continue

        result, num_vars, num_constrs = payload if payload is not None else (None, None, None)

        if result is None:
            print("No solution found (infeasible).")
        else:
            print_board(result["X"], "Click matrix X")
            print(f"Total clicks: {result['total_clicks']}")
            if result["T"] is not None:
                print_board(result["T"], "Hardest target T")
            ok = verify_solution(N, c,
                                 result["T"] if result["T"] else target,
                                 result["X"])
            print(f"Verification: {'PASS' if ok else 'FAIL'}")

        print(f"Time: {elapsed:.3f}s")
        export_to_excel(instance_name, sheet_name, result, elapsed, num_vars, num_constrs, path=xlsx_path)